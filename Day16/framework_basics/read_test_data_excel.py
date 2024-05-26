import openpyxl

xl_location = r"D:\SeleniumPython\Projects\FITA\pythonProject1_FITA\Day16\excel_login_data.xlsx"


wb1 = openpyxl.load_workbook(xl_location)

# check sheetname
print(wb1.sheetnames)
sh1 = wb1['Sheet']
# check max number of rows and columns which have data filled
max_row = sh1.max_row
max_col = sh1.max_column
print(max_row)
print(max_col)

# get data in excel
for row in range(1, max_row+1):
    for col in range(1, max_col+1):
        data = sh1.cell(row, col).value
        print(data)

# write/ update data in excel
sh1.cell(4, 1).value = "admin@yourstor3.com"
sh1.cell(4, 2).value = "admin3"

wb1.save(xl_location)



