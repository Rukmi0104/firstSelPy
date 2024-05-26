import openpyxl

# code to create excel file

wb = openpyxl.Workbook()  # --> with this new excel will be created
print(wb.sheetnames)  # Print sheet names in excel to check number of sheets o/p is in list form-> ['Sheet']

# save sheet with some name
file_location = r"D:\SeleniumPython\Projects\FITA\pythonProject1_FITA\Day16\excel_login_data.xlsx"
wb.save(file_location)

# fill login data in excel in Sheet
sh = wb['Sheet']
sh.cell(1, 1).value = "username"
sh.cell(1, 2).value = "password"

sh.cell(2, 1).value = "admin@yourstor1.com"
sh.cell(2, 2).value = "admin1"
sh.cell(3, 1).value = "admin@yourstore.com"
sh.cell(3, 2).value = "admin"

wb.save(file_location)
