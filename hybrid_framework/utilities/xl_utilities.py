import openpyxl


def get_max_row(xl_path, sheet):
    wb = openpyxl.load_workbook(xl_path)
    sh = wb[sheet]
    return sh.max_row


def get_max_col(xl_path, sheet):
    wb = openpyxl.load_workbook(xl_path)
    sh = wb[sheet]
    return sh.max_column


def read_cell(xl_path, sheet, r, c):
    wb = openpyxl.load_workbook(xl_path)
    sh = wb[sheet]
    return sh.cell(r, c).value


def write_cell(xl_path, sheet, r, c, data):
    wb = openpyxl.load_workbook(xl_path)
    sh = wb[sheet]
    sh.cell(r,c).value=data
    wb.save(xl_path)


